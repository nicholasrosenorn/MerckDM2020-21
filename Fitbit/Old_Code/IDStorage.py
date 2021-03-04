import pandas as pd
from openpyxl import load_workbook
import os


def inputID ():
    choice = str(input("Would you like to add a new user? [Y/N] "))
    # Check for user spellings
    if choice in ("yes", "Yes", "y", "YES"):
        choice = "Y"
    if choice in ("no", "NO", "n", "No"):
        choice = "N"
    
    # Input User Check
    while (choice != "Y") and (choice != "N"):
        choice = str(input("Input error. Please Try again. \nWould you like to add a new user? [Y/N] "))

    if choice == "Y":
        inputConfirmation = "N"
        while (inputConfirmation not in ("Y", "y", "YES", "yes", "Yes")):
            cID = str(input("Please input the client username: "))
            secretID = str(input("Please input client password: "))
            version = str(input("Please input your fitbit model: "))
            inputConfirmation = str(input("Please confirm the following information: \nUsername: %s\nPassowrd: %s\nModel: %s\n[Y/N] " %(cID, secretID, version)))
        excelAppend(cID,secretID,version)
    else:
        print("Have a good day! Be well!")

    return()

def excelAppend(CLIENT, SECRET, VERSION):
    
    try: # if Fitbit_IDs.xlsx already exists 
        # Locate file in same folder as code
        file_dir = os.path.dirname(os.path.abspath(__file__))
        read_path = file_dir + '/Fitbit_IDs.xlsx'

        # Create a new dataframe with the input information
        df1 = pd.DataFrame({'Username': [CLIENT], 'Password': [SECRET], 'Model_Version': [VERSION]})
    
        # Open the existing excel document
        book = load_workbook(read_path)
        
        # Prep the excel document to be appended
        writer = pd.ExcelWriter(read_path, engine='openpyxl')
        writer.book = book
        writer.sheets = {ws.title: ws for ws in book.worksheets}
        
        # Append the excel document
        df1.to_excel(writer,sheet_name='Sheet1', startrow=writer.sheets['Sheet1'].max_row, index = False,header= False)
        writer.save()
        print("File appended")
        

    except IOError: # Start a new userID data base
        print("No pre-existing document found. Creating a new one.")
        writer = pd.ExcelWriter('Fitbit_IDs.xlsx', engine = 'xlsxwriter')
        file_dir = os.path.dirname(os.path.abspath(__file__))
        file_path = os.path.join(file_dir, writer)
        df = pd.DataFrame({'Username': [CLIENT], 'Password': [SECRET], 'Model_Version': [VERSION]})
        df.to_excel(file_path, sheet_name='Sheet1', index=False)
        print("New file created")
    
    return()

inputID()